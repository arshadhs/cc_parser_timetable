import json
from typing import OrderedDict
from datetime import datetime
import requests
import xmltodict
from openpyxl import load_workbook, Workbook
from dateutil import parser

def get_prayer_table(year):
    url = 'https://www.moonsighting.com/praytable.php'
    parameters = {'year': str(year), 'tz' : 'Europe/London', 'lat': '52.2178,', 'lon': '0.0662', 'method': '0', 'both': 'false', 'time': '0'}
    response = requests.get(url, params=parameters)
    if response.status_code != 200:
        raise Exception(f"Failed to get prayer time table from {url}")
    start = response.text.find('<div')
    end = response.text.rfind('</div>')
    xml_text = response.text[start:end+len('</div>')]
    elements = xmltodict.parse(xml_text)
    header = elements['div']['table']['thead']['tr']['th']
    header[0] = 'Date'
    data = OrderedDict()
    data = {'header': header}
    schedule = OrderedDict()
    for day in elements['div']['table']['tbody']['tr']:
        date, Fajr, Sunrise, Dhuhr, Asr, Maghrib, Isha = day["td"]
        schedule[date] = OrderedDict(Fajr=Fajr, Sunrise=Sunrise, Dhuhr=Dhuhr, Asr=Asr, Maghrib=Maghrib, Isha=Isha)
    data['schedule'] = schedule
    return data

def _get_sheet_from_hdr(wb, headers):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=1, max_col=7, max_row=1):
            header = []

            for cell in row:
                # print (cell.value)
                # print (type(cell.value))
                #val = cell.value if not isinstance (cell.value, float) else int(cell.value) 
                header.append(str(cell.value))
            break               

        # print (headers)
        # print(header)
        if set(header).issuperset(headers):
#            print(f"Found xls '{sheet_name}'")
            return sheet
        else:
            print (set(header).difference(headers))
            print (set(headers).difference(header))

    print('Failed to find the xls with timing information')
    return None

def get_donation_sheet(filename, year):
    iwb = load_workbook(filename, read_only=True)
    return _get_sheet_from_hdr(iwb, {year, 'Fajr', 'Sunrise', 'Dhuhr', 'Asr(H)', 'Maghrib', 'Isha'})


def get_prayer_table_offline(year, filename):
    sheet = get_donation_sheet(filename, year)

#    print ("sheet: ", type(sheet))
    schedule = OrderedDict()

#    header[0] = 'Date'
    data = OrderedDict()
#    data = {'header': header}

    if sheet:
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):

            date_string, Fajr, Sunrise, Dhuhr, Asr, Maghrib, Isha = row[:7]
            #print (row)
            # Combine the year with the date string
            full_date_string = f"{date_string}"

            # Convert the string to a datetime object using strptime
            print(full_date_string)
            date_obj = (parser.parse(full_date_string)).date() #datetime.strptime(full_date_string, '%Y-%b-%d %H:%M:%S').date()
            #date_obj = datetime.strptime(date_string, '%Y %b %d %a').date()
            if date_obj is not None:
                schedule[date_obj] = OrderedDict(
                                        Fajr=datetime.strptime(str(Fajr), '%H:%M:%S').time(),
                                        Sunrise=datetime.strptime(str(Sunrise), '%H:%M:%S').time(),
                                        Dhuhr=datetime.strptime(str(Dhuhr), '%H:%M:%S').time(),
                                        Asr=datetime.strptime(str(Asr), '%H:%M:%S').time(),
                                        Maghrib=datetime.strptime(str(Maghrib), '%H:%M:%S').time(),
                                        Isha=datetime.strptime(str(Isha), '%H:%M:%S').time())

    # for key, value in (schedule).items():
        # print ("key: ", key)
        # print ("value: ", type(schedule))
        # print ("value: ", type(value))
        # print ("value: ", value)
        
        # for k, v in (value).items():
                # print ("key: ", k)
                # print ("value: ", type(v))
                # print ("value: ", v)

    data['schedule'] = schedule
    return data

