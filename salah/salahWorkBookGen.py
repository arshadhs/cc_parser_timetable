#!/usr/bin/env python

"""
    Workbook reading / writing
"""

__author__      = "Mohammad Azim Khan, Arshad H. Siddiqui"
__copyright__   = "Free to all"

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from salah import Salah

COLOUR_BLUE = "add8e6"
COLOR_P_BLUE = "1e7ba0"
COLOR_S_BLUE = "0a2842"

COLOUR_GREY = "dbdbdb"
COLOR_L_GREY = "6e6e6e"
COLOR_D_GREY = "474747"

class SalahWorkBook(object):
    def __init__(self, salah, usage): # name, date, start, details, has_jamat=True, fill_color=None, header_color=None):
        self.salah = salah
        self.usage = usage

        self.header_color = PatternFill("solid", fgColor=COLOR_L_GREY)
        self.fill_color = PatternFill("solid", fgColor=COLOUR_BLUE) if self.salah.is_juma else PatternFill("solid", fgColor=COLOUR_GREY)

    # __str__ method to customize how the object is printed
    def __str__(self):
        return f"SalahWorkBook(header_color={self.header_color}, fill_color={self.fill_color})"

    # Date format
    def displayTime(self, time):
        return time.strftime("%H:%M") if self.usage == "booking" else time.strftime("%H:%M:%S")

    def _style_header(self, ws, row, col):
        if self.header_color:
            ws.cell(row, col).fill = self.header_color
        ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    # Month - Header - Month, Year, Fajr, Dhuhr, Asr, Maghrib, Isha
    def add_xl_top_header(self, ws, row, col, hide, do_not_merge=False):
        start_col = col
        ws.cell(row, col).value = self.salah.name       # xlsx: Fajr, Dhuhr, Asr, Maghrib, Isha

        if (not hide):
            ws.cell(row, col).value = self.salah.name       # xlsx: Fajr, Dhuhr, Asr, Maghrib, Isha
        else:
            if (self.salah.name == "Fajr"):
                ws.cell(row, col).value = "Morning"
            if (self.salah.name == "Dhuhr"):
                ws.cell(row, col).value = "Noon"
            if (self.salah.name == "Asr"):
                ws.cell(row, col).value = "After Noon"
            if (self.salah.name == "Maghrib"):
                ws.cell(row, col).value = "Evening"
            if (self.salah.name == "Isha"):
                ws.cell(row, col).value = "Night"

        self._style_header(ws, row, col)
        col += 1
        if self.salah.has_jamat:
            col += 3                                    # xlsx: Jamat, booking and location - add columns
        if not do_not_merge:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=col - 1)
        return col

    # Month - Sub Header - Start, Booking, Jamat, Location
    def add_xl_header(self, ws, row, col):
        ws.cell(row, col).value = 'Start' if self.usage == "booking" else self.salah.name.lower()                 # Start - Header
        self._style_header(ws, row, col)
        col += 1
#        if self.salah.has_jamat:
        if self.usage == "booking":
            ws.cell(row, col).value = 'Booking'# Booking - Header
            self._style_header(ws, row, col)
            col += 1

        ws.cell(row, col).value = 'Jamat' if self.usage == "booking" else self.salah.name.lower() + "_con"    # Jamat - Header
        self._style_header(ws, row, col)
        col += 1
        ws.cell(row, col).value = 'Location' if self.usage == "booking" else self.salah.name.lower() + "_loc"  # Location - Header
        self._style_header(ws, row, col)
        col += 1
        return col

    # Values (we.cell(row,col)) - Start, Booking, Jamat, Location (Fill the value and style / colour etc.)
    def add_xl_values(self, ws, row, col):
        fill_color = PatternFill("solid", fgColor=COLOUR_BLUE) if self.salah.is_juma else PatternFill("solid", fgColor=COLOUR_GREY) if row % 2 != 0 else PatternFill(fgColor="ffffff")

        # Start
        ws.cell(row, col).value = self.displayTime(self.salah.start)  # Start Time
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if self.salah.is_juma else Font(bold=False)
        col += 1

        # If there's congregation - Booking, Jamat, Location
#        if self.salah.has_jamat:
        # Booking
        if self.usage == "booking":
            ws.cell(row, col).value = self.displayTime(self.salah.booking_start) + " - " + self.displayTime(self.salah.booking_end) if self.salah.booking_start else ""
            ws.cell(row, col).fill = fill_color
            ws.cell(row, col).font = Font(bold=True) if self.salah.is_juma else Font(bold=False)
            col += 1

        # Jamat
        ws.cell(row, col).value = self.displayTime(self.salah.jamat) if self.salah.jamat else "" # Jamat
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if self.salah.is_juma else Font(bold=False)
        col += 1

        # Location
        ws.cell(row, col).value = self.salah.location
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if self.salah.is_juma else Font(bold=False)
        col += 1

        return col


# To add Sunrise column
class FajrSalahWorkBook(SalahWorkBook):
    def __init__(self, salah, usage):
        super().__init__(salah, usage)
        self.sunrise = self.salah.details['Sunrise'] #details['Sunrise']
        self.usage = usage

    def add_xl_top_header(self, ws, row, col, hide):
        start_col = col

        col = super().add_xl_top_header(ws, row, col, hide, do_not_merge=True)
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=col)
        col += 1
        return col

    def add_xl_header(self, ws, row, col):
        col = super().add_xl_header(ws, row, col)
        ws.cell(row, col).value = 'Sunrise' if self.usage == "booking" else "sunrise"
        self._style_header(ws, row, col)
        col += 1
        return col

    # Fill the we.cell(row,col). value with sunrise time
    def add_xl_values(self, ws, row, col):
        fill_color = PatternFill("solid", fgColor=COLOUR_BLUE) if self.salah.is_juma else PatternFill("solid", fgColor=COLOUR_GREY) if row % 2 != 0 else PatternFill(fgColor="ffffff")
        col = super().add_xl_values(ws, row, col)
        ws.cell(row, col).value = super().displayTime(self.sunrise)
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if self.salah.is_juma else Font(bold=False)
        col += 1
        return col
