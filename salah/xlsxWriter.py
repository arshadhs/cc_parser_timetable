#!/usr/bin/env python

"""
    xlsx writer
"""

__author__      = "Arshad H. Siddiqui"
__copyright__   = "Free to all"

import datetime
import argparse
import math

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

import salahUtils

from moonsighting import get_prayer_table, get_prayer_table_offline
from salah import Salah, recalculate_jamat_time
from salahWorkBookGen import SalahWorkBook, FajrSalahWorkBook
from validator import validateJamatTime

COLOUR_BLUE = "add8e6"
# COLOR_P_BLUE = "1e7ba0"
# COLOR_S_BLUE = "0a2842"

COLOUR_GREY = "dbdbdb"
COLOR_L_GREY = "6e6e6e"
# COLOR_D_GREY = "474747"

def writer(wbTable, year, usage, hideColumns):
    wb = Workbook()
    wb['Sheet'].title = 'CC booking'
    ws = wb['CC booking']

    header_color = PatternFill("solid", fgColor=COLOR_L_GREY)
    row = 1

    # Add each day as rows
    for serial_no, (date, day) in enumerate(wbTable['schedule'].items(), start=1):

        week_day = date.strftime('%a')
        is_juma = week_day == 'Fri'

        # Add Salah header

        # Header - web (2 rows every month)
        if (date.day == 1) and (date.month == 1) and usage == "web":
            # id, date, day
            col = 1
            ws.cell(row, col).value = 'id'
            ws.cell(row, col).fill = header_color
            ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
            col += 1
            ws.cell(row, col).value = 'date'
            ws.cell(row, col).fill = header_color
            ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
            col += 1

            # Start, Booking, Jamat, Location (Sunrise)
            for salah in next(iter(wbTable['schedule'].values())).values():
                if isinstance(salah, SalahWorkBook):
                    col = salah.add_xl_header(ws, row, col)
            row += 1

        # Header - booking (1 row on top of the sheet)
        elif int(date.strftime('%d') == "01") and usage == "booking":
            # First day of the month

            # Add top header
            col = 4
            ws.cell(row, 1).value = '{} {}'.format(date.strftime('%B'), date.strftime('%y'))    # Month and Year
            ws.cell(row, 1).font = Font(bold=True, color='FFFFFF')
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, 1).fill = header_color
            ws.cell(row, 2).fill = header_color
            ws.cell(row, 3).fill = header_color
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col - 1)

            # Fajr, Dhuhr, Asr, Maghrib, Isha
            for salah in next(iter(wbTable['schedule'].values())).values():
                if isinstance(salah, SalahWorkBook):
                    col = salah.add_xl_top_header(ws, row, col, hideColumns)

            # Add Salah header

            # id, date, day
            row += 1
            col = 1
            ws.cell(row, col).value = 'id'
            ws.cell(row, col).fill = header_color
            ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
            col += 1
            ws.cell(row, col).value = 'date'
            ws.cell(row, col).fill = header_color
            ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
            col += 1

            ws.cell(row, col).value = 'day'
            ws.cell(row, col).fill = header_color
            ws.cell(row, col).font = Font(bold=True, color='FFFFFF')
            col += 1

            # Start, Booking, Jamat, Location (Sunrise)
            for salah in next(iter(wbTable['schedule'].values())).values():
                if isinstance(salah, SalahWorkBook):
                    col = salah.add_xl_header(ws, row, col)
            row += 1

        # Data - fill up the actual data
        fill_color = PatternFill("solid", fgColor=COLOUR_BLUE) if is_juma else PatternFill("solid", fgColor=COLOUR_GREY) if row % 2 != 0 else PatternFill(fgColor="ffffff")
        col = 1
        ws.cell(row, col).value = serial_no                     # id (Serial Number)
        serial_no += 1
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if is_juma else Font(bold=False)

        col += 1                                                # Date
        ws.cell(row, col).value = date.strftime('%b-%d') if usage == "booking" else date.strftime('%d/%m/%Y')
        ws.cell(row, col).fill = fill_color
        ws.cell(row, col).font = Font(bold=True) if is_juma else Font(bold=False)

        if usage == "booking":                                  # Day
            col += 1
            ws.cell(row, col).value = date.strftime('%a')
            ws.cell(row, col).fill = fill_color
            ws.cell(row, col).font = Font(bold=True) if is_juma else Font(bold=False)

        col += 1                                                # Values - Start, Booking, Jamat, Location (Fill the value and style / colour etc.)
        for salah in day.values():
            if isinstance(salah, SalahWorkBook):
                col = salah.add_xl_values(ws, row, col)
        row += 1

    setCellWidth(ws)

    # Hide columns not relevant for booking
    if (hideColumns):
        for col in ['D', 'F', 'H', 'I', 'K', 'I', 'M', 'N', 'P', 'R', 'T']:
            ws.column_dimensions[col].hidden= True

        thin_border = Border(left=Side(style='none'), 
                     right=Side(style='thick'), 
                     top=Side(style='none'), 
                     bottom=Side(style='none'))

        for col in ['C', 'G', 'L', 'Q', 'U']:
            for cell in ws[col]:
                cell.border = thin_border
            #ws.cell(column=2).border = thin_border

    outFile = 'Cambourne_salah_timetable_'+year+'.xlsx'
    wb.save(outFile)
    print("\nWritten to", outFile)
    # if not_in_use(outFile):
        # wb.save(outFile)
        # print("\nWritten to", outFile)
    # else:
        # print("\nError[13]: Permission denied", outFile)


def not_in_use(filename):
        try:
            os.rename(filename,filename)
            return True
        except:
            return False


def setCellWidth(ws):
    dims = {}
    for row in ws.rows:
        numOfCol = 0
        for cell in row:
            numOfCol = 5
            from openpyxl.cell import MergedCell
            if cell.value and not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal='center')
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 5
