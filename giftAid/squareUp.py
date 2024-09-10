#!/usr/bin/env python

"""gift_aid.py: It takes an xl sheet input with donations and outputs
   output.xlsx with Gift Aid summary.
   
   Donations sheet is detected if it has a header containing:
    'Date', 'Value', 'Balance', 'Account Name', 'Account Number'

   Donor name is extracted as first string in comma separated Description text.
   Rows with negative or 0 donation are filtered out.
   More filters can be added for tidy output.
   Better to use another input of gift aiders.
   
   All donations from same donor are added and the last donation date is
   picked for summary output. Example output:
Title    First name    Last name    House name or number    Postcode    Aggregate donation    Sponsered event    Donation date   Amount
         A             A L I                                                                                     01-Feb-2022     10
         A             K h a n                                                                                   15-Oct-2021     40
         P             M U H A M M A D                                                                           10-May-2021     13
         R             S i d d i q u i                                                                           21-Jul-2021     10

"""

__author__      = "Mohammad Azim Khan"
__copyright__   = "Free to all"

import sys
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

########################################################################

def is_donor(data):
    if float(data[3]) > 0:
        return True
    # Add more filters to refine output
    #if len(desc) == 4: # Assuming interestig donors have this format
    return False

########################################################################

def _get_sheet_from_hdr(wb, headers):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_col=10, max_row=1):
            header = []
            #print(row)
            for cell in row:
                #print(cell)
                header.append(cell.value)
            break
        print(headers)
        print (header)
        if set(header).issuperset(headers):
            print(f"Found Statement '{sheet_name}'")
            return sheet
        else:
            print(headers.difference(header))
    print('Failed to find a sheet with donations information')
    return None

def get_SquareUp_sheet(wb):
    return _get_sheet_from_hdr(wb, {'Date', 'Category', 'Item', 'Gross Sales', 'Customer Name'})

def get_giftaid_sheet(wb):
    return _get_sheet_from_hdr(wb, {'First Name', 'Surname', 'House Number', 'Postcode', 'Account Name'})

########################################################################

def process_SquareUp(file_SquareUp, start, end):
    print(f"Opening file '{file_SquareUp}'")
    iwb = load_workbook(file_SquareUp, read_only=True)
    donors = {}
    sheet = get_SquareUp_sheet(iwb)
    if sheet:
        for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
#           print (row)
            timestamp, payment_type, desc, amount, ac_name = row[:5]
#            print(datetime.datetime.fromisoformat(str(timestamp)))
            
            # Date should fall in range
            entry_time = datetime.date.fromisoformat(str(timestamp).split(' ')[0])
            if entry_time < start or entry_time > end:
                continue
##            desc = [x.strip() for x in desc.split(',')]

            if not is_donor(row[:7]):
                continue
##            donor = desc[0]
#            print ("donor", donor)

            # remove ' from the start of Description - Donor
##            if (donor[0] == "'"): donor = donor[1:]

            # Add a new DONOR to list, or add the amount to already existing DONOR
            donor = ac_name
            print (donor)
            if donor not in donors:
                donors[donor] = {'total': float(amount), 'last donation': timestamp}
            else:
                donors[donor]['total'] += float(amount)
                if timestamp > donors[donor]['last donation']:
                    donors[donor]['last donation'] = timestamp
    return donors

# Read GA for Name, Address & Reference
def process_giftaid_data(file_GA):
    print(f"Opening file '{file_GA}'")
    iwb = load_workbook(file_GA, read_only=True)
    donors = {}
    sheet = get_giftaid_sheet(iwb)
    if sheet:
        gift_aid_data = {}
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            fname, surname, hno, postcode, ga, ac_name = row[:6]
            print("row", row)
            print("ac_name=", ac_name)
            if (fname != None):                
                gift_aid_data[ac_name.lower()] = {'fname':fname, 'surname': surname, 'hno': hno, 'postcode': postcode, 'ga' : ga}
                print("gift_aid_data=", gift_aid_data[ac_name.lower()])

        return gift_aid_data
    return None

########################################################################
# donors is a dictionary ----------- A DURRANI {'total': 336.83, 'last donation': datetime.datetime(2022, 12, 12, 0, 0)}
# gift_aid_data is a dictionary --   a durrani {'fname': 'Abdul Jabbar', 'surname': 'Durrani', 'hno': 5, 'postcode': 'CB23 6LF', 'ga': None}
#######################
# abdulm and alsal abdulm and alsal
# {'fname': 'Naim', 'surname': 'Abdulmohdi', 'hno': 76, 'postcode': 'CB23 5ed', 'ga': None} {'total': 255.0, 'last donation': datetime.datetime(2023, 2, 14, 0, 0)}
# abdulmohdi n abdulmohdi n
# {'fname': 'Naim', 'surname': 'Abdulmohdi', 'hno': 76, 'postcode': 'CB23 5ed', 'ga': None} {'total': 120.0, 'last donation': datetime.datetime(2023, 3, 10, 0, 0)}
#######################
def merge_donors(donors, gift_aid_data):
    print(f"++++++++++++++++++++++++++++++")
    print(f"merge_donors file")
#    print(donors)
#   print(gift_aid_data)
    
    mergedData = {}
    d = {}
    
    for i, (k, v) in enumerate(gift_aid_data.items()):
#        print(i, k, v)
        
        for a, (b, c) in enumerate(donors.items()):

            if(k == b.lower()):
                print(a, b, c)
#               print (k, b.lower())
                #print(v, c)
                x = (v['fname'], v['surname'], v['hno'], v['postcode'] , v['ga'])
                if x in mergedData:
#                    print(mergedData[x]['total'])
#                    print(c['total'])
                    total = mergedData[x]['total'] + c['total']
                    mergedData[x].update({'total': total})
                    
                    if mergedData[x]['last donation'] < c['last donation']:
 #                       print(mergedData[x]['last donation'])
 #                       print(c['last donation'])
                        timestamp = c['last donation']
                        mergedData[x].update({'last donation': timestamp})                  

                    mergedData[x]["ref" + str(i)] = k
               
 #                   print(mergedData[x])
                else:
                    mergedData[x] = {'total': c['total'], 'last donation': c['last donation'], "ref" + str(i): k}
 #                   print(mergedData[x])
               
    for i, (k, v) in enumerate(mergedData.items()):
       print(i, k, v)
    
    return mergedData

########################################################################

def write_merge_donors(mergedData):
    print(f"++++++++++++++++++++++++++++++")
    print(f"write_merge_donors file")
#    print(f"write_merge_donors file '{mergedData}'")
    wb = Workbook()
    wb['Sheet'].title = 'Gift Aid'
    ws = wb['Gift Aid']

    # Write header
    for col, x in enumerate(("Title", "First name", "Last name",
                             "House name or number","Postcode",
                             "Aggregate donation", "Sponsered event",
                             "Donation date", "Amount", "Reference", "GA"), start=1):
            ws.cell(1, col).value = x
            ws.cell(1, col).font = Font(bold=True, color='FFFFFF')
            ws.cell(1, col).fill = PatternFill("solid", fgColor="000000")
    
    # Thought, we can have a first pass on data to estimate column widths
    
    for row, (donor, donation) in enumerate(mergedData.items(), start=2):
        print (donor, donation)
#        print (type(donor))
#        print(type(donation))
#        print("FIRST", donor.split(' ')[:-1])
#        print("SECOND", donor.split(' ')[-1])
        # donor_data = gift_aid_data.get(donor.lower(), {
            # 'fname': ' '.join(donor.split(' ')[:-1]),
            # "surname": ''.join(donor.split(' ')[-1]),
            # "hno": '',
            # 'postcode': '',
            # 'ga': ''})
        title = ''
        # first_name = donor_data['fname'] #
# #        print("first", first_name)
        # last_name = donor_data['surname'] #
# #        print ("last_name", last_name)
        # address = donor_data['hno']
        # postcode = donor_data['postcode']
        aggregate_donation = ''
        event = ''
        # donation_date = donation['last donation']
        # amount = donation['total']
        # ga = donor_data['ga']
        
        for col, x in enumerate((title, donor[0], donor[1], donor[2], donor[3], aggregate_donation, event, donation['last donation'], donation['total'], "", donor[4],), start=1):
            if isinstance(x, datetime.datetime):
                ws.cell(row, col).value = x.strftime('%d-%b-%Y')
                ws.cell(row, col).alignment = Alignment(horizontal='right')
            else:
                ws.cell(row, col).value = x
    print("saving git aid summary in output.xlsx")    
    wb.save('output.xlsx')
    
def write_donor_summary(donors, gift_aid_data):
    print(f"++++++++++++++++++++++++++++++")
    print(f"write_donor_summary file '{gift_aid_data}'")
    wb = Workbook()
    wb['Sheet'].title = 'Gift Aid'
    ws = wb['Gift Aid']

    # Write header
    for col, x in enumerate(("Title", "First name", "Last name",
                             "House name or number","Postcode",
                             "Aggregate donation", "Sponsered event",
                             "Donation date", "Amount", "Reference", "GA"), start=1):
            ws.cell(1, col).value = x
            ws.cell(1, col).font = Font(bold=True, color='FFFFFF')
            ws.cell(1, col).fill = PatternFill("solid", fgColor="000000")
    
    # Thought, we can have a first pass on data to estimate column widths
    
    print (gift_aid_data)
    for row, (donor, donation) in enumerate(donors.items(), start=2):
        print (donor)
#        print("FIRST", donor.split(' ')[:-1])
#        print("SECOND", donor.split(' ')[-1])
        donor_data = gift_aid_data.get(donor.lower(), {
            'fname': ' '.join(donor.split(' ')[:-1]),
            "surname": ''.join(donor.split(' ')[-1]),
            "hno": '',
            'postcode': '',
            'ga': ''})
        title = ''
        first_name = donor_data['fname'] #
#        print("first", first_name)
        last_name = donor_data['surname'] #
#        print ("last_name", last_name)
        address = donor_data['hno']
        postcode = donor_data['postcode']
        aggregate_donation = ''
        event = ''
        donation_date = donation['last donation']
        amount = donation['total']
        ga = donor_data['ga']
        
        for col, x in enumerate((title, first_name, last_name, address, postcode, aggregate_donation, event, donation_date, amount, donor, ga), start=1):
            if isinstance(x, datetime.datetime):
                ws.cell(row, col).value = x.strftime('%d-%b-%Y')
                ws.cell(row, col).alignment = Alignment(horizontal='right')
            else:
                ws.cell(row, col).value = x
    print("Saving Git aid summary in output.xlsx")    
    wb.save('output.xlsx')

########################################################################

# C:\GA>squareUp.py items_SquareUp-2022-23.xlsx GA.xlsx 2022

if __name__ == '__main__':
    file_SquareUp, file_GA, _start = sys.argv[1:]
    start = datetime.date(int(_start), 4, 5)
    end = datetime.date(int(_start) + 1, 4, 4)
    gift_aid_data = process_giftaid_data(file_GA)
    donors = process_SquareUp(file_SquareUp, start, end)
    #write_donor_summary(donors, gift_aid_data)
    
    mergedData = merge_donors(donors, gift_aid_data)
    write_merge_donors(mergedData)

########################################################################



            # timestamp, payment_type, desc, value, balance, ac_name, ac_num = row[:7]
            # desc = [x.strip() for x in desc.split(',')]
            # if not is_donor(row[:7]):
            #     continue
            # donor = desc[0]
            # if donor not in donors:
            #     donors[donor] = {'total': float(value), 'last donation': timestamp}
            # else:
            #     donors[donor]['total'] += float(value)
            #     if timestamp > donors[donor]['last donation']:
            #         donors[donor]['last donation'] = timestamp