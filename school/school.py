#!/usr/bin/env python

"""school.py:
    Input: 
        (1) Bank Statement - donations xlsx sheet,  
        (2) Reference Sheet, xlsx with names and references to search for
        (3) Start year to search
    Output: output.xlsx with monthly School Fee records - from Sept to August

   Bank Statement sheet is detected if it has a header containing:
    'Date', 'Value', 'Balance', 'Account Name', 'Account Number'

   Donor name is extracted as first string in comma separated Description text.
   Rows with negative or 0 donation are filtered out.
   More filters can be added for tidy output.

   Reference Sheet - use as input for Parent/Donor Name and reference

   All donations from same donor are added for the month. Example output:
        Parent's Name	Reference	September	October	November	December	January	February	March	April	May	June	July	August
        Abdellatif Salah	a salah	120	200	110	0	0	0	0	0	0	0	0	0
        Abdul Durrani	a durrani	75	75	75	0	0	0	0	0	0	0	0	0
        Adli Mohamed A Alghanimi	n elbaruni	40	40	40	0	0	0	0	0	0	0	0	0

    C:\GA>school.py NatWest-download-2022-23-School.xlsx school.xlsx 2022

"""

__author__ = "AHS"
__copyright__ = "Free to all"

import sys
import datetime
import argparse

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

########################################################################

def is_donor(data):
    if float(data[3]) > 0:
        return True
    # Add more filters to refine output
    #if len(desc) == 4: # Assuming interesting donors have this format
    return False

########################################################################

def _get_sheet_from_hdr(wb, headers):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_col=10, max_row=1):
            header = []

            for cell in row:

                header.append(cell.value)
            break

        if set(header).issuperset(headers):
            print(f"Found Bank Statement '{sheet_name}'")
            return sheet

#    print('Failed to find Bank Statement with donations information')
    return None

def get_donation_sheet(wb):
    return _get_sheet_from_hdr(wb, {'Date', 'Type', 'Description', 'Value', 'Balance', 'Account Name', 'Account Number'})

def get_reference_sheet(wb):
    return _get_sheet_from_hdr(wb, {'Parent', 'Account Name'})

import csv

def convert_csv__toxlsx(filename):
    wb = wb = Workbook()
    ws = wb.active

    with open(filename) as f:
        reader = csv.reader(f, delimiter=':')
        for row in reader:
            ws.append(row)

    wb.save('account.xlsx')

########################################################################

# Bank Statement
def process_bank_st(filename, start, end):
    print(process_bank_st.__name__ + " === START ===")
    print(f"Opening file '{filename}'")
    
    # if filename.endswith('.csv'):
        # convert_csv__toxlsx(filename)
        # filename = 'account.xlsx'
    
    iwb = load_workbook(filename, read_only=True)
    donors = {}
    sheet = get_donation_sheet(iwb)
    if sheet:
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            timestamp, payment_type, desc, value, balance, ac_name, ac_num = row[:7]
            #print(datetime.datetime.fromisoformat(str(timestamp)))
            
            # Date should fall in range
            entry_time = datetime.date.fromisoformat(str(timestamp).split(' ')[0])
            if entry_time < start or entry_time > end:
                continue
            desc = [x.strip() for x in desc.split(',')]

            if not is_donor(row[:7]):
                continue
            donor = desc[0]
 #           print ("donor", donor)

            # remove ' from the start of Description - Donor
            if (donor[0] == "'"): donor = donor[1:]
#           print ("donor", donor)

            # Add a new DONOR to list, or add the amount to already existing DONOR
            if donor not in donors:
                donors[donor] = {'total'+str(entry_time.month): float(value)}#, 'last donation': timestamp}
            else:
                if('total'+str(entry_time.month) in donors[donor]):
                    donors[donor]['total'+str(entry_time.month)] += float(value)
                else:
                    donors[donor]['total'+str(entry_time.month)] = float(value)
                # if timestamp > donors[donor]['last donation']:
                    # donors[donor]['last donation'] = timestamp
                #print("Donor: ", donor)
    print(process_bank_st.__name__ + " === END ===")
    return donors

# Parent's Name & Reference
def process_reference_data(filename):
    print(process_reference_data.__name__ + " === START ===")
    print(f"Opening file '{filename}'")
    iwb = load_workbook(filename, read_only=True)
    sheet = get_reference_sheet(iwb)
    if sheet:
        reference_data = {}
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[0] is not None:
#                print (row)
                fname, ac_name, numOfChildren = row[:3]

                reference_data[ac_name.lower()] = {'fname':fname, 'numOfChildren': int(numOfChildren)}
            else:
                break

        print(process_reference_data.__name__ + " === END 1 ===")
        return reference_data
    print(process_reference_data.__name__ + " === END 2 ===")
    return None

########################################################################

# Write to an xlsx
def write_output(donors, reference_data, outPath):
    print(write_output.__name__ + " === START ===")
    wb = Workbook()
    wb['Sheet'].title = 'School Fee'
    ws = wb['School Fee']
    total = 0

    # Write header
    for col, x in enumerate(("Parent's Name", "Reference",
                            "September", "October", "November", "December", "January", "February", "March", "April", "May", "June", "July", "August",
                            "Child(ren)", "Paid", "Due"), start=1):
            ws.cell(1, col).value = x
            ws.cell(1, col).font = Font(bold=True, color='FFFFFF')
            ws.cell(1, col).fill = PatternFill("solid", fgColor="000000")
    
    # Thought, we can have a first pass on data to estimate column widths
    row = 2 # row 1 is header

    output = {}
    list_of_lists = [] 
    
    # Iterates over Name and Reference (data from reference sheet)
    for reference, ac_name in reference_data.items():
#        print("reference_data ", reference, ac_name)
        fee_data = donors.get(reference.upper())
#        print("fee_data: ", fee_data)
        first_name = ac_name['fname'] #
              
#       reference = reference
        total1 = fee_data.get('total1', 0) if (fee_data != None) else 0
        total2 = fee_data.get('total2', 0) if (fee_data != None) else 0
        total3 = fee_data.get('total3', 0) if (fee_data != None) else 0
        total4 = fee_data.get('total4', 0) if (fee_data != None) else 0
        total5 = fee_data.get('total5', 0) if (fee_data != None) else 0
        total6 = fee_data.get('total6', 0) if (fee_data != None) else 0
        total7 = fee_data.get('total7', 0) if (fee_data != None) else 0
        total8 = fee_data.get('total8', 0) if (fee_data != None) else 0
        total9 = fee_data.get('total9', 0) if (fee_data != None) else 0
        total10 = fee_data.get('total10', 0) if (fee_data != None) else 0
        total11 = fee_data.get('total11', 0) if (fee_data != None) else 0
        total12 = fee_data.get('total12', 0) if (fee_data != None) else 0

        total = total + total1 + total2 + total3 + total4 + total5 + total6 + total7 + total8 + total9 + total10 + total11 + total12

        # If already the fname exists, catenate reference
        if first_name not in output:
            output[first_name] = {'reference' : reference}
#            print("    output[first_name].get('reference') :: " + output[first_name].get('reference'))
        else:
            output[first_name] = {'reference' : output[first_name].get('reference') + " + " + reference}
#            print("output[first_name].get('reference') = " + output[first_name].get('reference'))

        position = next((i for i, row in enumerate(list_of_lists) if row[0] == first_name), None)     
        
        if position is not None:
            list_of_lists[position] = [first_name, output[first_name].get('reference'), 
                                        list_of_lists[position][2]+total9,
                                        list_of_lists[position][3]+total10,
                                        list_of_lists[position][4]+total11,
                                        list_of_lists[position][5]+total12, 
                                        list_of_lists[position][6]+total1,
                                        list_of_lists[position][7]+total2,
                                        list_of_lists[position][8]+total3,
                                        list_of_lists[position][9]+total4, 
                                        list_of_lists[position][10]+total5,
                                        list_of_lists[position][11]+total6,
                                        list_of_lists[position][12]+total7,
                                        list_of_lists[position][13]+total8,
                                        ac_name['numOfChildren'], 0, 0
                                       ]
        else:
            masterdata = [first_name, output[first_name].get('reference'), total9, total10, total11, total12, total1, total2, total3, total4, total5, total6, total7, total8, ac_name['numOfChildren'], 0, 0]
            list_of_lists.append(masterdata)

    # Output to xlsx
    for r in list_of_lists: 
        column = 0
        total = 0
        due = 0
#        print(r)
        for element in r:
            column += 1          
            
            if (column > 2 and column < 15):
                total += element
            if column == 15:
                numOfChildren = element
                ws.cell(row, column).value = element
            if column == 16:
                ws.cell(row, column).value = total
            elif column == 17:               
                amountToBePAid = 480 if numOfChildren == 1 else 900 if numOfChildren == 2 else 1080
                due = amountToBePAid - total                
                ws.cell(row, column).value = due if due > 0 else 0
                if (due > 0):
                    ws.cell(row, column).font = Font(bold=True, color='000000')
                    ws.cell(row, column).fill = PatternFill("solid", fgColor="D3D3D3")

            else:
                ws.cell(row, column).value = element
                if (element == 0):
                        ws.cell(row, column).font = Font(bold=True, color='000000')
                        ws.cell(row, column).fill = PatternFill("solid", fgColor="D3D3D3")

        row = row+1

    #print (output)
#    print("list_of_lists: ", list_of_lists)
    print("Saving School Fee summary in output.xlsx")
    outFileName = 'output.xlsx' if outPath is None else outPath + "\\" 'output.xlsx'
    wb.save(outFileName)
    print(write_output.__name__ + " === END ===")

########################################################################

def main():
    parser = argparse.ArgumentParser(description='Process school arguments.')
    parser.add_argument('--bank', dest='bankStatementXLS', default=None, help='Bank Statement')
    parser.add_argument('--reference', dest='referenceXLS', default=None, help='Reference Statement')
    parser.add_argument('--year', dest='year', default=datetime.datetime.now().year - 1, help='Year of school records')

    parser.add_argument('--outputPath', dest='outPath', help='XLS file for timetable')

    args = parser.parse_args()

    bankStatementXLS = args.bankStatementXLS
    referenceXLS = args.referenceXLS
    year = args.year
    outPath = args.outPath

#    bankStatementXLS, referenceXLS, _start = sys.argv[1:]

    start = datetime.date(int(year), 9, 1)#4, 5)
    end = datetime.date(int(year) + 1, 8, 31)#4, 4)
    print(start)
    print(end)
    
    reference_data = process_reference_data(referenceXLS)
    donors = process_bank_st(bankStatementXLS, start, end)

    write_output(donors, reference_data, outPath)
    

# C:\GA>school.py NatWest-download-2022-23-School.xlsx school.xlsx 2022
if __name__ == '__main__':
    main()

########################################################################
